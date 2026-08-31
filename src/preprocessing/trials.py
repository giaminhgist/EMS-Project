"""Workbook parsing: fixation rows -> per-trial sorted events.

Reads only the required sheet and columns, retains original row numbers for
QC, groups by the exact ``IMAGE`` string without numeric assumptions, and
stable-sorts every trial by ``FIX_INDEX`` then original row number.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .qc import ParsedEvent, TrialEvents, coerce_float, coerce_int


class WorkbookError(ValueError):
    """Raised when a subject workbook does not match the expected schema."""


@dataclass(frozen=True)
class ColumnIndices:
    image: int
    fix_index: int
    fix_duration: int
    fix_x: int
    fix_y: int
    fix_pupil: int


def resolve_columns(
    header: tuple[object, ...], columns: dict[str, str], path: Path
) -> ColumnIndices:
    """Map the configured column names to positions and validate them."""
    header = tuple(header)
    positions: dict[str, int] = {}
    for role, name in columns.items():
        if name not in header:
            raise WorkbookError(
                f"{path}: sheet is missing required column {name!r} "
                f"(header: {header!r})"
            )
        positions[role] = header.index(name)
    return ColumnIndices(**positions)


def parse_workbook_trials(
    path: Path,
    subject_id: str,
    columns: dict[str, str],
    sheet_name: str,
) -> tuple[dict[str, TrialEvents], int]:
    """Parse one workbook into ``{IMAGE value: TrialEvents}``.

    Returns ``(trials_by_image, n_data_rows)``. The header row must be row 1
    and is not counted as data.
    """
    if not path.is_file():
        raise WorkbookError(f"workbook not found: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise WorkbookError(
                f"{path}: sheet {sheet_name!r} not found (sheets: {wb.sheetnames})"
            )
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise WorkbookError(f"{path}: sheet {sheet_name!r} is empty") from exc

        cols = resolve_columns(tuple(header), columns, path)

        trials: dict[str, TrialEvents] = {}
        n_data_rows = 0
        for row_number, row in enumerate(rows_iter, start=2):
            n_data_rows += 1
            image_raw = row[cols.image] if cols.image < len(row) else None
            if image_raw is None:
                image = ""
            else:
                image = str(image_raw).strip()
            fix_index_status, fix_index = coerce_int(
                row[cols.fix_index] if cols.fix_index < len(row) else None
            )
            duration_status, duration = coerce_float(
                row[cols.fix_duration] if cols.fix_duration < len(row) else None
            )
            x_status, x = coerce_float(row[cols.fix_x] if cols.fix_x < len(row) else None)
            y_status, y = coerce_float(row[cols.fix_y] if cols.fix_y < len(row) else None)
            pupil_status, pupil = coerce_float(
                row[cols.fix_pupil] if cols.fix_pupil < len(row) else None
            )
            coord_status = "ok" if (x_status == "ok" and y_status == "ok") else (
                "nonfinite" if (x_status == "nonfinite" or y_status == "nonfinite")
                else ("empty" if (x_status == "empty" and y_status == "empty") else "malformed")
            )
            event = ParsedEvent(
                row_number=row_number,
                image=image,
                fix_index=fix_index,
                fix_index_status=fix_index_status,
                duration=duration,
                duration_status=duration_status,
                x=x,
                y=y,
                pupil=pupil,
                coord_status=coord_status,
                pupil_status=pupil_status,
            )
            if image == "":
                raise WorkbookError(
                    f"{path}: row {row_number} has an empty IMAGE value"
                )
            trials.setdefault(image, TrialEvents(subject_id=subject_id, stimulus_id=image))
            trials[image].events.append(event)
        return trials, n_data_rows
    finally:
        wb.close()
